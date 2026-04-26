"""xlsx parser."""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from hks.errors import ExitCode, KSError
from hks.ingest.fingerprint import ParserFlags
from hks.ingest.models import ParsedDocument
from hks.ingest.office_common import (
    Segment,
    SkippedSegment,
    build_placeholder,
    merge_skipped,
    to_markdown_table,
)


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _append_package_placeholders(
    path: Path, segments: list[Segment], skipped_segments: list[SkippedSegment]
) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
    except (OSError, zipfile.BadZipFile) as exc:
        raise KSError(
            f"corrupt file: {path}",
            exit_code=ExitCode.DATAERR,
            code="CORRUPT",
            details=[str(exc)],
        ) from exc

    if any(name.startswith("xl/charts/") for name in names):
        merge_skipped(skipped_segments, SkippedSegment(type="chart"))
        segments.append(Segment(kind="placeholder", text=build_placeholder("chart")))
    if any("pivot" in name.lower() for name in names):
        merge_skipped(skipped_segments, SkippedSegment(type="pivot"))
        segments.append(Segment(kind="placeholder", text=build_placeholder("pivot")))
    if any(name.endswith("vbaProject.bin") for name in names):
        merge_skipped(skipped_segments, SkippedSegment(type="macros"))
        segments.append(Segment(kind="placeholder", text=build_placeholder("macros")))
    if any(name.startswith("xl/embeddings/") for name in names):
        merge_skipped(skipped_segments, SkippedSegment(type="embedded_object"))
        segments.append(
            Segment(
                kind="placeholder",
                text=build_placeholder("embedded_object", "ole"),
            )
        )


def parse(path: Path, flags: ParserFlags) -> ParsedDocument:
    del flags
    try:
        value_book = load_workbook(path, read_only=True, data_only=True)
        formula_book = load_workbook(path, read_only=True, data_only=False)
    except (InvalidFileException, ValueError, KeyError, OSError, zipfile.BadZipFile) as exc:
        raise KSError(
            f"corrupt file: {path}",
            exit_code=ExitCode.DATAERR,
            code="CORRUPT",
            details=[str(exc)],
        ) from exc

    segments: list[Segment] = []
    skipped_segments: list[SkippedSegment] = []
    for sheet_name in value_book.sheetnames:
        value_sheet = value_book[sheet_name]
        formula_sheet = formula_book[sheet_name]
        segments.append(
            Segment(
                kind="sheet_header",
                text=f"## {sheet_name}",
                metadata={"sheet_name": sheet_name},
            )
        )

        header: list[str] | None = None
        for row_index, (formula_row, value_row) in enumerate(
            zip(formula_sheet.iter_rows(), value_sheet.iter_rows(), strict=False),
            start=1,
        ):
            cell_pairs = list(zip(formula_row, value_row, strict=False))
            row_values: list[str] = []
            formula_only = False
            for formula_cell, value_cell in cell_pairs:
                if getattr(formula_cell, "data_type", None) == "f":
                    if value_cell.value is None:
                        row_values.append(_stringify(formula_cell.value))
                        formula_only = True
                    else:
                        row_values.append(_stringify(value_cell.value))
                else:
                    row_values.append(_stringify(value_cell.value))

            if header is None:
                header = [cell or f"col_{index + 1}" for index, cell in enumerate(row_values)]
                continue
            if not header or not any(value.strip() for value in row_values):
                continue
            row_meta: dict[str, Any] = {
                "sheet_name": sheet_name,
                "row_index": row_index,
                "header_row": header,
            }
            if formula_only:
                row_meta["formula_only"] = True
            segments.append(
                Segment(
                    kind="table_row",
                    text=to_markdown_table(header, [row_values]),
                    metadata=row_meta,
                )
            )

    _append_package_placeholders(path, segments, skipped_segments)
    return ParsedDocument(
        title=path.stem.replace("-", " ").replace("_", " ").strip() or path.stem,
        body="",
        format="xlsx",
        segments=segments,
        skipped_segments=skipped_segments,
    )
